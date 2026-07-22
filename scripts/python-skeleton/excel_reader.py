"""
Excel Data Reader Module

Handles reading and processing Excel files for the ACMG SF Variants Report.
Maps Excel data to template variables according to the data mapping guide.

BioVarFlow_HemOnc branch: also handles the optional HemOnc sheets
('HemOnc (P-LP)', 'HemOnc Coverage gaps', 'HemOnc Genes Coverage') written
by generate_lean_report_org.py when --hemonc-genes is supplied.
"""

import pandas as pd
from datetime import datetime
from typing import Dict, List, Any, Optional


class ExcelDataReader:
    """Reads and processes Excel data for report generation"""

    def __init__(self, excel_path: str, debug: bool = False):
        self.excel_path = excel_path
        self.debug = debug
        self.required_tabs = {
            'Sample Summary': True,
            'ACMG SF (P-LP)': True,
            'Coverage gaps': False,  # Optional
            'ACMG Genes Coverage': False,  # Optional
            # HemOnc sheets — optional; present only on runs that supplied
            # --hemonc-genes to the lean-report script.
            'HemOnc (P-LP)': False,
            'HemOnc Coverage gaps': False,
            'HemOnc Genes Coverage': False,
        }

    def read_all_tabs(self) -> Dict[str, pd.DataFrame]:
        """Read all relevant tabs from Excel file"""
        if self.debug:
            print(f"Reading Excel file: {self.excel_path}")

        try:
            # Read the Excel file
            import openpyxl
            excel_data = pd.read_excel(self.excel_path, sheet_name=None)

            if self.debug:
                print(f"Available tabs: {list(excel_data.keys())}")

            # Check for required tabs
            for tab_name, required in self.required_tabs.items():
                if tab_name not in excel_data and required:
                    raise ValueError(f"Required tab '{tab_name}' not found in Excel file")

            # Special handling for ClinVar_Link column to extract URLs from
            # HYPERLINK formulas. Applied to every variants sheet that carries
            # such formulas (ACMG SF and, on the HemOnc branch, HemOnc too).
            for variants_sheet in ('ACMG SF (P-LP)', 'HemOnc (P-LP)'):
                if variants_sheet in excel_data:
                    excel_data[variants_sheet] = self._extract_hyperlinks(
                        self.excel_path,
                        variants_sheet,
                        excel_data[variants_sheet],
                    )

            return excel_data

        except Exception as e:
            raise Exception(f"Error reading Excel file: {e}")

    def _extract_hyperlinks(self, excel_path: str, sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
        """Extract URLs from Excel HYPERLINK formulas"""
        import openpyxl
        import re
        
        # Check if ClinVar_Link column exists
        if 'ClinVar_Link' not in df.columns:
            return df
        
        try:
            # Load workbook with openpyxl to access formulas
            wb = openpyxl.load_workbook(excel_path, data_only=False)
            ws = wb[sheet_name]
            
            # Find column index for ClinVar_Link
            header_row = [cell.value for cell in ws[1]]
            if 'ClinVar_Link' not in header_row:
                return df
                
            col_idx = header_row.index('ClinVar_Link') + 1
            
            # Extract URLs from HYPERLINK formulas
            urls = []
            for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                cell = ws.cell(row=row, column=col_idx)
                cell_value = cell.value
                
                if cell_value and isinstance(cell_value, str) and cell_value.startswith('=HYPERLINK'):
                    # Extract URL from formula: =HYPERLINK("URL","Display")
                    match = re.search(r'=HYPERLINK\("([^"]+)"', cell_value)
                    if match:
                        urls.append(match.group(1))
                    else:
                        urls.append(None)
                elif cell_value:
                    urls.append(str(cell_value))
                else:
                    urls.append(None)
            
            # Update the DataFrame with extracted URLs
            df['ClinVar_Link'] = urls
            
            if self.debug:
                print(f"Extracted {sum(1 for u in urls if u)} ClinVar links from HYPERLINK formulas")
            
        except Exception as e:
            if self.debug:
                print(f"Warning: Could not extract hyperlinks: {e}")
        
        return df

    def validate_data(self, data: Dict[str, pd.DataFrame]) -> None:
        """Validate that required columns exist in each tab"""
        validations = {
            'Sample Summary': ['Sample_ID'],  # Add other required columns
            'ACMG SF (P-LP)': ['Gene', 'HGVSc', 'HGVSp', 'ClinVar'],
        }

        for tab_name, required_columns in validations.items():
            if tab_name in data:
                df = data[tab_name]
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    raise ValueError(f"Missing required columns in '{tab_name}': {missing_columns}")

    def process_for_template(self, data: Dict[str, pd.DataFrame], sample_id: str, assay: Optional[str] = None) -> Dict[str, Any]:
        """Process Excel data into format needed for HTML template"""

        template_data = {
            'sample_id': sample_id,
            'report_generated': datetime.now().strftime('%Y-%m-%d %H:%M'),
        }

        # Add assay if provided (overrides Excel data)
        if assay:
            template_data['assay'] = assay

        # Process Page 1 header data
        template_data.update(self._process_header_data(data))

        # Override assay from command line if provided
        if assay:
            template_data['assay'] = assay

        # Process variant data for Page 1 & 2
        template_data.update(self._process_variant_data(data))

        # Process QC metrics for Page 3
        template_data.update(self._process_qc_metrics(data))

        # Process coverage data for Page 3
        template_data.update(self._process_coverage_data(data))

        # Process HemOnc data (BioVarFlow_HemOnc branch). Only fills the
        # dedicated HemOnc pages when the corresponding sheets are present in
        # the workbook; otherwise the HemOnc pages render with their default
        # "no data" placeholders (safe on ACMG-only runs).
        template_data.update(self._process_hemonc_data(data))

        return template_data

    def _process_header_data(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Extract header information from Sample Summary tab"""
        if 'Sample Summary' not in data:
            return {}

        df = data['Sample Summary']
        header_data = {}

        # Map Excel columns to template variables based on actual lean report structure
        column_mapping = {
            'Sample_ID': 'sample_id',
            'Assay': 'assay',
            'Build': 'reference_build',
        }

        # Extract first row of data (assuming single sample per file)
        if len(df) > 0:
            for excel_col, template_var in column_mapping.items():
                if excel_col in df.columns:
                    header_data[template_var] = df[excel_col].iloc[0]

        # Set defaults for missing data
        header_data.setdefault('assay', 'WES')
        header_data.setdefault('reference_build', 'GRCh38')
        header_data.setdefault('pipeline', 'BioVarFlow_HemOnc v1.1.0')
        header_data.setdefault('databases', 'ClinVar (2025-01), gnomAD v4.1, Ensembl VEP Release 115, REVEL (latest release), AlphaMissense (Science 2023, updated 2025-05)')

        return header_data

    def _process_variant_data(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Extract and filter variant data for Page 1 and 2"""
        if 'ACMG SF (P-LP)' not in data:
            return {
                'page1_variants': [],
                'page2_variants': []
            }

        df = data['ACMG SF (P-LP)']
        # Filter for pathogenic/likely pathogenic variants
        pathogenic_mask = df['ClinVar'].isin(['Pathogenic', 'Likely pathogenic','Conflicting_classifications_of_pathogenicity'])
        pathogenic_variants = df[pathogenic_mask]

        # Split based on star rating (>=2 stars for Page 1, <2 stars for Page 2)
        page1_variants = []
        page2_variants = []

        for _, row in pathogenic_variants.iterrows():
            # Get ClinVar link - now extracted from HYPERLINK formula in read_all_tabs
            clinvar_id = row.get('ClinVar_Link', '')
            
            # If still no link, try to construct from HGVSc as fallback
            if not clinvar_id or pd.isna(clinvar_id):
                if pd.notna(row.get('HGVSc')):
                    hgvsc = str(row.get('HGVSc', ''))
                    if ':' in hgvsc:
                        # Extract just the HGVS notation after the colon
                        hgvs_notation = hgvsc.split(':')[-1]
                        clinvar_id = f"https://www.ncbi.nlm.nih.gov/clinvar/?term={quote_plus(hgvs_notation)}"
                else:
                    clinvar_id = ''
            
            variant_data = {
                'gene': row.get('Gene', '—'),
                'hgvsc': row.get('HGVSc', '—'),
                'hgvsp': row.get('HGVSp', '—'),
                'clinvar_id': clinvar_id if clinvar_id else '',
            }


            # Check star rating
            stars = row.get('ClinVar_Stars', 0)
            try:
                stars = int(stars) if pd.notna(stars) else 0
            except (ValueError, TypeError):
                stars = 0

            if stars >= 2:
                page1_variants.append(variant_data)
            else:
                page2_variants.append(variant_data)

        return {
            'page1_variants': page1_variants,
            'page2_variants': page2_variants
        }

    def _process_qc_metrics(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Extract QC metrics from Sample Summary tab"""
        if 'Sample Summary' not in data:
            return {}

        df = data['Sample Summary']
        print(df)
        metrics = {}

        # Map Excel columns to template variables based on actual lean report structure
        if len(df) > 0:
            row = df.iloc[0]
            
            # Handle coverage metrics
            mean_cov = row.get('Mean_Coverage', '—')
            if mean_cov != '—' and pd.notna(mean_cov):
                metrics['mean_coverage'] = f"{mean_cov:.2f}x"
            else:
                metrics['mean_coverage'] = '—'
            
            # Handle mapping percentage
            mapped_pct = row.get('Mapped_Percent', '—')
            if mapped_pct != '—' and pd.notna(mapped_pct):
                metrics['mapped_percent'] = f"{mapped_pct:.2f}%"
            else:
                metrics['mapped_percent'] = '—'
            
            # Handle duplicate percentage
            dup_pct = row.get('Duplicate_Percent', '—')
            if dup_pct != '—' and pd.notna(dup_pct):
                metrics['duplicate_percent'] = f"{dup_pct:.2f}%"
            else:
                metrics['duplicate_percent'] = '—'
            
            # Handle insert size
            insert_size = row.get('Mean_Insert_Size', '—')
            if insert_size != '—' and pd.notna(insert_size):
                metrics['mean_insert_size'] = f"{insert_size:.1f}bp"
            else:
                metrics['mean_insert_size'] = '—'
            
            # Handle ACMG coverage - this might be in a different sheet or calculated
            # For now, set a default or try to find it in coverage data
            metrics['acmg_covered_percent'] = row.get('ACMG_PctRegionsCovered', '—')  # Will be updated from coverage data if available

        return metrics

    def _process_coverage_data(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Extract coverage data for both coverage gaps and overall coverage sections"""
        coverage_data = {
            'coverage_gaps': [],
            'overall_coverage': []
        }

        # Process coverage gaps from "Coverage gaps" sheet
        if 'Coverage gaps' in data:
            gaps_df = data['Coverage gaps']
            for _, row in gaps_df.iterrows():
                coverage_pct = row.get('Pct>=20x', 100)
                try:
                    coverage_pct = float(coverage_pct) if pd.notna(coverage_pct) else 100
                except (ValueError, TypeError):
                    coverage_pct = 100

                gene_data = {
                    'gene': row.get('Gene', '—'),
                    #'transcript': row.get('MANE_ID', '—'),
                    'ExonStart': row.get('ExonStart', '—'),
                    'ExonEnd': row.get('ExonEnd', '—'),
                    'coverage': f"{coverage_pct:.1f}%"
                }

                # Add to gaps if coverage < 100%
                if coverage_pct < 100:
                    coverage_data['coverage_gaps'].append(gene_data)

        # Process overall coverage from "ACMG SF Genes Coverage" sheet
        if 'ACMG SF Genes Coverage' not in data:
            return coverage_data

        df = data['ACMG SF Genes Coverage']

        # Process overall coverage (length-weighted mean coverage per gene)
        for _, row in df.iterrows():
            coverage_pct = row.get('Pct20', 100)  # This is the length-weighted mean
            try:
                coverage_pct = float(coverage_pct) if pd.notna(coverage_pct) else 100
            except (ValueError, TypeError):
                coverage_pct = 100

            gene_data = {
                'gene': row.get('Gene', '—'),
                'transcript': row.get('MANE_ID', '—'),
                'coverage': f"{coverage_pct:.1f}%"
            }

            # Add all genes to overall coverage
            coverage_data['overall_coverage'].append(gene_data)

        return coverage_data

    # ---------------------------------------------------------------
    # HemOnc (BioVarFlow_HemOnc branch)
    # ---------------------------------------------------------------
    def _process_hemonc_data(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """
        Extract HemOnc-scoped data from the workbook.

        Returns a dict with keys the template processor consumes:
          - hemonc_page1_variants     : list[dict] — variants with ClinVar ≥2 stars
          - hemonc_page2_variants     : list[dict] — variants with ClinVar <2 stars / VUS
          - hemonc_coverage_gaps      : list[dict] — HemOnc exons with %20x < 100
          - hemonc_overall_coverage   : list[dict] — per-gene HemOnc coverage rollup
          - hemonc_covered_percent    : str        — Sample Summary metric formatted %
          - hemonc_available          : bool       — True iff any HemOnc sheet was present

        All values default to empty / '—' when the corresponding sheets are
        absent, so this method is safe to call on ACMG-only workbooks.
        """
        result: Dict[str, Any] = {
            'hemonc_page1_variants': [],
            'hemonc_page2_variants': [],
            'hemonc_coverage_gaps': [],
            'hemonc_overall_coverage': [],
            'hemonc_covered_percent': '—',
            'hemonc_available': False,
        }

        # ---- variants (mirrors _process_variant_data's ACMG logic) ----
        if 'HemOnc (P-LP)' in data:
            result['hemonc_available'] = True
            df = data['HemOnc (P-LP)']
            if 'ClinVar' in df.columns and len(df) > 0:
                pathogenic_mask = df['ClinVar'].isin([
                    'Pathogenic',
                    'Likely pathogenic',
                    'Conflicting_classifications_of_pathogenicity',
                ])
                pathogenic_variants = df[pathogenic_mask]

                for _, row in pathogenic_variants.iterrows():
                    clinvar_id = row.get('ClinVar_Link', '')
                    if not clinvar_id or pd.isna(clinvar_id):
                        if pd.notna(row.get('HGVSc')):
                            hgvsc = str(row.get('HGVSc', ''))
                            if ':' in hgvsc:
                                from urllib.parse import quote_plus
                                hgvs_notation = hgvsc.split(':')[-1]
                                clinvar_id = (
                                    "https://www.ncbi.nlm.nih.gov/clinvar/?term="
                                    + quote_plus(hgvs_notation)
                                )
                        else:
                            clinvar_id = ''

                    variant_data = {
                        'gene':  row.get('Gene', '—'),
                        'hgvsc': row.get('HGVSc', '—'),
                        'hgvsp': row.get('HGVSp', '—'),
                        'clinvar_id': clinvar_id if clinvar_id else '',
                    }

                    stars = row.get('ClinVar_Stars', 0)
                    try:
                        stars = int(stars) if pd.notna(stars) else 0
                    except (ValueError, TypeError):
                        stars = 0

                    if stars >= 2:
                        result['hemonc_page1_variants'].append(variant_data)
                    else:
                        result['hemonc_page2_variants'].append(variant_data)

        # ---- coverage gaps (mirrors ACMG "Coverage gaps" handling) ----
        if 'HemOnc Coverage gaps' in data:
            result['hemonc_available'] = True
            gaps_df = data['HemOnc Coverage gaps']
            for _, row in gaps_df.iterrows():
                coverage_pct = row.get('Pct>=20x', 100)
                try:
                    coverage_pct = float(coverage_pct) if pd.notna(coverage_pct) else 100
                except (ValueError, TypeError):
                    coverage_pct = 100

                gene_data = {
                    'gene':      row.get('Gene', '—'),
                    'ExonStart': row.get('ExonStart', '—'),
                    'ExonEnd':   row.get('ExonEnd', '—'),
                    'coverage':  f"{coverage_pct:.1f}%",
                }

                if coverage_pct < 100:
                    result['hemonc_coverage_gaps'].append(gene_data)

        # ---- per-gene overall coverage grid ----
        # Unmeasured genes (Pct20 = NaN) are rendered as "NA" so the grid
        # remains honest about scope. This is important when the pipeline
        # was run against a BED that doesn't include every HemOnc gene
        # (e.g. a legacy ACMG-only mosdepth run followed by a HemOnc-aware
        # lean-report re-render): the 79 non-overlap HemOnc genes still
        # appear on the report, clearly marked as not measured.
        if 'HemOnc Genes Coverage' in data:
            result['hemonc_available'] = True
            df = data['HemOnc Genes Coverage']
            for _, row in df.iterrows():
                raw_pct = row.get('Pct20', None)
                if raw_pct is None or pd.isna(raw_pct):
                    coverage_str = 'NA'
                else:
                    try:
                        coverage_str = f"{float(raw_pct):.1f}%"
                    except (ValueError, TypeError):
                        coverage_str = 'NA'

                result['hemonc_overall_coverage'].append({
                    'gene':       row.get('Gene', '—'),
                    'transcript': row.get('MANE_ID', '—'),
                    'coverage':   coverage_str,
                })

        # ---- HemOnc-scoped Sample Summary metric ----
        if 'Sample Summary' in data and len(data['Sample Summary']) > 0:
            ss_row = data['Sample Summary'].iloc[0]
            pct = ss_row.get('HemOnc_PctRegionsCovered', None)
            if pct is not None and pd.notna(pct):
                try:
                    result['hemonc_covered_percent'] = f"{float(pct):.2f}%"
                except (ValueError, TypeError):
                    result['hemonc_covered_percent'] = str(pct)

        return result